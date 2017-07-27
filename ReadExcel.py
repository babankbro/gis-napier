import openpyxl

def readProvince(filepath):
    wb = openpyxl.load_workbook(filepath)
    sheet =wb.get_active_sheet()


    Header = {}
    indexHeader = {}
    row = 1
    col = 1
    while True:
        cell = sheet.cell(row = row, column=col)
        if cell.value == None:
            break
        name = cell.value;
        Header[name] = col
        indexHeader[col] = name
        col+=1
    print(Header)

    colAmpher = Header["อำเภอของฟาร์ม"]
    ampherNames = {}
    row = 2
    while True:
        cell = sheet.cell(row = row, column=colAmpher)
        ampherName = cell.value;
        if ampherName == None:
            break
        
        if ampherName in ampherNames:
            ampherNames[ampherName] += 1
        else:
            ampherNames[ampherName] = 1

        row +=1
    numberOfRow = row
    print(ampherNames)

    firstCol = Header["รวมโคเนื้อ"]
    tables = {}
    for ampherName in ampherNames:
        tables[ampherName] = {}
        for col in range(firstCol, len(Header)+1):
            tables[ampherName][indexHeader[col]] = 0
    #print(tables)

    for name in tables:
        print(name)

    for row in range(2, numberOfRow+1):
        ampherName = sheet.cell(row = row, column=colAmpher).value
        if ampherName not in tables:
            print(ampherName)
            continue
        ampherInfo = tables[ampherName]
        for col in range(firstCol, len(Header)+1):
            value = int(sheet.cell(row = row, column=col).value)
            ampherInfo[indexHeader[col]] += value
    return tables

def display(tables, columnNames = None):
    if columnNames is None:
        columnNames = []
        for key in tables:
            for name in tables[key]:
                columnNames.append(name)
            break
    result = "\t\t\t"
    for ampherName in tables:
        result += ampherName + "\t"
    result += "\n"
    m = 19
    for name in columnNames:
        if len(name) <= 14:
            result += name + "\t\t\t"
        elif len(name) <= 22:
            result += name + "\t\t"
        else:
            result += name + "\t"
        for ampherName in tables:
            result += str(tables[ampherName][name]) + "\t"
        result += "\n"
    print(result)




"""animal = "รวมโคเนื้อ"
totalValue = 0;
for name in tables:
    print(name, tables[name][animal])
    totalValue += tables[name][animal]
print(animal, totalValue)
"""



